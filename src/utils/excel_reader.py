"""Excel test data reader built with OpenPyXL."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelReader:
    """Read Excel workbooks and return test data as dictionaries."""

    def __init__(self, file_path: str | Path, data_only: bool = True) -> None:
        self.file_path = Path(file_path).resolve()
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file was not found: {self.file_path}")
        self.workbook = load_workbook(filename=self.file_path, data_only=data_only)

    def get_sheet(self, sheet_name: str) -> Worksheet:
        if sheet_name not in self.workbook.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' was not found in workbook: {self.file_path}")
        return self.workbook[sheet_name]

    def get_sheet_names(self) -> list[str]:
        return list(self.workbook.sheetnames)

    def get_cell_value(self, sheet_name: str, row: int, column: int | str) -> Any:
        sheet = self.get_sheet(sheet_name)
        if isinstance(column, str):
            return sheet[f"{column}{row}"].value
        return sheet.cell(row=row, column=column).value

    def get_row_data(self, sheet_name: str, row_number: int) -> dict[str, Any]:
        sheet = self.get_sheet(sheet_name)
        headers = self._get_headers(sheet)
        values = [cell.value for cell in sheet[row_number]]
        return self._map_row_to_dict(headers, values)

    def get_all_rows(self, sheet_name: str) -> list[dict[str, Any]]:
        sheet = self.get_sheet(sheet_name)
        headers = self._get_headers(sheet)
        rows: list[dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(value is not None for value in row):
                rows.append(self._map_row_to_dict(headers, list(row)))
        return rows

    def get_column_data(self, sheet_name: str, column: int | str) -> dict[str, list[Any]]:
        sheet = self.get_sheet(sheet_name)
        column_index = self._column_letter_to_index(column) if isinstance(column, str) else column
        header = sheet.cell(row=1, column=column_index).value
        values = [
            sheet.cell(row=row_number, column=column_index).value
            for row_number in range(2, sheet.max_row + 1)
        ]
        return {str(header): [value for value in values if value is not None]}

    def get_sheet_data(self, sheet_name: str) -> list[dict[str, Any]]:
        return self.get_all_rows(sheet_name)

    def get_workbook_data(self) -> dict[str, list[dict[str, Any]]]:
        return {sheet_name: self.get_all_rows(sheet_name) for sheet_name in self.get_sheet_names()}

    def close(self) -> None:
        self.workbook.close()

    @staticmethod
    def _get_headers(sheet: Worksheet) -> list[str]:
        headers = [cell.value for cell in sheet[1]]
        if not headers or all(header is None for header in headers):
            raise ValueError(f"Sheet '{sheet.title}' must contain headers in the first row.")
        return [str(header).strip() if header is not None else f"column_{index}" for index, header in enumerate(headers, 1)]

    @staticmethod
    def _map_row_to_dict(headers: list[str], values: list[Any]) -> dict[str, Any]:
        return {header: values[index] if index < len(values) else None for index, header in enumerate(headers)}

    @staticmethod
    def _column_letter_to_index(column: str) -> int:
        column = column.strip().upper()
        if not column.isalpha():
            raise ValueError(f"Column must be a letter or number, got: {column}")
        index = 0
        for character in column:
            index = index * 26 + ord(character) - ord("A") + 1
        return index
